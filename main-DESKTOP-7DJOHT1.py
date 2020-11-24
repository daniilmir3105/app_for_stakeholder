import base_class_analys
from PyQt5 import QtCore, QtGui, QtWidgets
import sys
from interface_2 import Ui_Dialog
import base_class_analys
import become_result
import os
import xlwt
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd 

# making objects for analys
analys_param_1 = base_class_analys.impact_level()
analys_param_2 = base_class_analys.engagement()
analys_param_3 = base_class_analys.needs_and_requirements()
analys_param_4 = base_class_analys.expectations()
analys_param_5 = base_class_analys.level_of_interests()
analys_param_6 = base_class_analys.effect()
analys_param_7 = base_class_analys.problematical_character()
analys_param_8 = base_class_analys.communicativity()

# making object to become a result 
final_result = become_result.result()

# Dialog = QtWidgets.QDialog()
# ui = Ui_Dialog()
# ui.setupUi(Dialog)
# Dialog.show()

class Stakeholder(QtWidgets.QMainWindow, Ui_Dialog):
    '''
    This class will have all fields and methods for stakeholder analys
    '''

    __result = []

    def __init__(self):
        '''
        Construktor of our class
        '''
        super().__init__()

        # Создание формы и Ui (наш дизайн)
        self.setupUi(self)

        # Показать наше окно
        self.show()

        self.pushButton.clicked.connect(self.making_analys)

    def get_data(self):
        '''
        Method, thar return some data of person
        '''

        return self.__result
        
    def making_analys(self):
        '''
        This function will make analys and get the result 
        '''
        # score result of parametrs 
        name_line = self.lineEdit.text()
        family_name_line = self.lineEdit_2.text()
        organisation_name_line = self.lineEdit_3.text()   
        position_line = self.lineEdit_4.text()
        contacts_line = self.lineEdit_5.text()

        # data that we become from comboboxes(our main-parametrs for analys)
        param_1 = self.comboBox.currentText()
        param_2 = self.comboBox_2.currentText()
        param_3 = self.comboBox_3.currentText()
        param_4 = self.comboBox_4.currentText()
        param_5 = self.comboBox_5.currentText()
        param_6 = self.comboBox_6.currentText()
        param_7 = self.comboBox_7.currentText()
        param_8 = self.comboBox_8.currentText()

        self.__result.append(name_line)
        self.__result.append(family_name_line)
        self.__result.append(organisation_name_line)
        self.__result.append(position_line)
        self.__result.append(contacts_line)
        self.__result.append(param_1)
        self.__result.append(param_2)
        self.__result.append(param_3)
        self.__result.append(param_4)
        self.__result.append(param_5)
        self.__result.append(param_6)
        self.__result.append(param_7)
        self.__result.append(param_8)
        g_1 = analys_param_1.make_analys(self.__result[5])
        g_2 = analys_param_2.make_analys(self.__result[6])
        g_3 = analys_param_3.make_analys(self.__result[7])
        g_4 = analys_param_4.make_analys(self.__result[8])
        g_5 = analys_param_5.make_analys(self.__result[9])
        g_6 = analys_param_6.make_analys(self.__result[10])
        g_7 = analys_param_7.make_analys(self.__result[11])
        g_8 = analys_param_8.make_analys(self.__result[12])

        # final result for matrix of stakeholder
        result_x = final_result.scoring_1(g_1=g_1, g_2=g_2, g_7=g_7, g_8=g_8)
        result_y = final_result.scoring_2(g_3=g_3, g_4=g_4, g_5=g_5, g_6=g_6)

        # what to do with this person
        final = final_result.make_simple_analys(x=result_x, y=result_y)

        if final == 'Данная личность не является стейкхолдером.':
            final_data = os.startfile('.\not.png')
        elif final == 'Данный стейкхолдер относится к блоку «Хорошие отношения»:\nC этими стейкхолдерами необходимо установить тесные рабочие отношения, потому что для них проект важен, они вовлечены в реализацию и активно влияют на процесс и результат.':
            final_data = os.startfile('.\good_relation.png')
        elif final == 'Данный стейкхолдер относится к блоку «Мониторинг»:\nОни имеют власть над реализацией проекта, но не слишком заинтересованы в нем.\nПри таком сочетании факторов они могут стать источниками рисков, поэтому необходимы тщательный мониторинг и внимательный менеджмент.':
            final_data = os.startfile('.\monitor.png')
        elif final == 'Данный стейкхолдер относится к блоку «Низкий приоритет»: этот стейкхолдер вовлечен и относительно заинтересован, но от него зависит не так много, поэтому с точки зрения распределения менеджерского внимания, у него низкий приоритет.':
            final_data = os.startfile('.\low_priority.png')
        else: 
            final_data = os.startfile('.\protect.png')
    
        if os.path.exists(r'.\database.xlsx') == True:
            # list of strings

            # Calling DataFrame constructor on list
            # df = pd.DataFrame(arr)

            wb = load_workbook('database.xlsx')
            ws = wb.active

            # for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(self.__result)

            wb.save('database.xlsx')
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(['Имя', 'Фамилия', 'Организация', 'Должность', 'Контактные данные', 'Уровень влияния', 'Заинтересованность', 'Нужды и потребности', 'Ожидания', 'Уровень заинтересованности', 'Эффект на ПМ', 'Проблематичность', 'Коммуникативность', 'Советы'])
            ws.append(self.__result)
            wb.save("database.xlsx")

        self.__result = []

        return final_data

if __name__ == "__main__":
    # Новый экземпляр QApplication
    app = QtWidgets.QApplication(sys.argv)
    
    # Сoздание инстанса класса, который мы создадим далее
    stakeholder = Stakeholder()
    # stakeholder.get_data()
    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog)
    
    sys.exit(app.exec_())
