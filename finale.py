from PyQt5 import QtCore, QtGui, QtWidgets
import sys
from interface_2 import Ui_Dialog
# import base_class_analys
import become_result
import os
import xlwt
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from abc import ABCMeta, abstractmethod


class base_analys_of_parametr(metaclass=ABCMeta):
    '''
    This metaclass will have merhods
    '''

    #
    # __parametr = 0

    # this will be base-method for getting the vallue of parametr
    @abstractmethod
    def make_analys(self):
        pass


class impact_level(base_analys_of_parametr):
    '''
    In this class will be our method for analys of impact from stakeholder
    '''

    __parametr_1 = 0

    def make_analys(self, impact_level) -> base_analys_of_parametr:
        '''
        This is a method, that will score parametr of impact on the project
        '''

        self.__parametr_1 = 2 if impact_level == 'высокий' else -2 if impact_level == 'низкий' else 0
        return self.__parametr_1


class engagement(base_analys_of_parametr):
    '''
    In this class will be our method for analys of engagement from stakeholder
    '''

    __parametr_2 = 0

    def make_analys(self, engagement_rate) -> base_analys_of_parametr:
        '''
        This is a method, that will score parametr of engagement of stakeholder
        '''

        self.__parametr_2 = 2 if engagement_rate == 'вовлеченность' else 1 if engagement_rate == 'лояльность' else -1 if engagement_rate == 'нелояльность' else -2 if engagement_rate == 'вредительство' else 0
        return self.__parametr_2


class needs_and_requirements(base_analys_of_parametr):
    '''
    In this class will be our method for analys of needs and requirements from stakeholder
    '''

    __parametr_3 = 0

    def make_analys(self, needs_requirements) -> base_analys_of_parametr:
        '''
        This is a method, that will score parametr of needs and requirements of stakeholder
        '''

        self.__parametr_3 = 2 if needs_requirements == 'прибыль' else 1 if needs_requirements == 'паблисити' else -2 if needs_requirements == 'провал проекта' else 0
        return self.__parametr_3


class expectations(base_analys_of_parametr):
    '''
    In this class will be our method for analys of expectations of stakeholder
    '''

    __parametr_4 = 0

    def make_analys(self, expectations) -> base_analys_of_parametr:
        '''
        This is a method, that will score parametr of expectations about project from stakeholder
        '''

        self.__parametr_4 = 2 if expectations == 'успех' else -2 if expectations == 'неудача' else 0
        return self.__parametr_4


class level_of_interests(base_analys_of_parametr):
    '''
    In this class will be our method for analys of level of interests of stakeholder
    '''

    __parametr_5 = 0

    def make_analys(self, level_and_quality_of_interest) -> base_analys_of_parametr:
        '''
        This is a method, that will score parametr of level of interests of stakeholder
        '''

        self.__parametr_5 = 2 if level_and_quality_of_interest == 'союзники' else 1 if level_and_quality_of_interest == 'поддерживающие' else -1 if level_and_quality_of_interest == 'неохотно участвующие' else -2 if level_and_quality_of_interest == 'оппоненты' else 0
        return self.__parametr_5


class effect(base_analys_of_parametr):
    '''
    In this class will be our method for analys of effect on project manager of stakeholder
    '''

    __parametr_6 = 0

    def make_analys(self, effect_of_a_pm_on_a_stakeholder) -> base_analys_of_parametr:
        '''
        This is a method, that will score parametr of level of effect on project manager of stakeholder
        '''

        self.__parametr_6 = 2 if effect_of_a_pm_on_a_stakeholder == 'высокий' else 1 if effect_of_a_pm_on_a_stakeholder == 'средний' else -2 if effect_of_a_pm_on_a_stakeholder == 'низкий' else 0
        return self.__parametr_6


class problematical_character(base_analys_of_parametr):
    '''
    This class will show how problematic is a stakeholder
    '''

    __parametr_7 = 0

    def make_analys(self, problem) -> base_analys_of_parametr:
        '''
        This method will show how problematic is a stakeholder
        '''

        self.__parametr_7 = 2 if problem == 'безпроблемный' else -1 if problem == 'средне-проблематичный' else -2 if problem == 'проблематичный' else 0
        return self.__parametr_7


class communicativity(base_analys_of_parametr):
    '''
    This class will show how communicative is a stakeholder
    '''

    __parametr_8 = 0

    def make_analys(self, communicative) -> base_analys_of_parametr:
        '''
        This method will show how communicative is a stakeholder
        '''

        self.__parametr_8 = 2 if communicative == 'высокая' else -2 if communicative == 'низкая' else 0
        return self.__parametr_8


# from abc import ABC, abstractmethod

class result:
    '''
    Class, that makes mathematic operations for become a result:
    '''

    def scoring_1(self, g_1, g_2, g_7, g_8):
        ''' Score the results. '''

        x = g_8 + g_2 + g_1 + g_7

        return x

    def scoring_2(self, g_3, g_4, g_5, g_6):
        ''' Score the results. '''

        y = g_3 + g_4 + g_5 + g_6

        return y

    def make_simple_analys(self, x, y):
        ''' Making simple analys for stakeholder '''

        if x == 0 and y == 0:
            result = 'Данная личность не является стейкхолдером.'
            return result
        elif x > 0 and y > 0:
            result = 'Данный стейкхолдер относится к блоку «Хорошие отношения»:\nC этими стейкхолдерами необходимо установить тесные рабочие отношения, потому что для них проект важен, они вовлечены в реализацию и активно влияют на процесс и результат.'
            return result
        elif x > 0 and y < 0:
            result = 'Данный стейкхолдер относится к блоку «Мониторинг»:\nЗдесь на карту нанесены стейкхолдеры, которые имеют власть над реализацией проекта, но не слишком заинтересованы в нем.\nПри таком сочетании факторов они могут стать источниками рисков, поэтому тут требуются тщательный мониторинг и внимательный менеджмент.'
            return result
        elif x < 0 and y < 0:
            result = 'Данный стейкхолдер относится к блоку «Низкий приоритет»: этот стейкхолдер вовлечен и относительно заинтересован, но от него зависит не так много, поэтому с точки зрения распределения менеджерского внимания, у него низкий приоритет.'
            return result
        else:
            result = 'Данный стейкхолдер относится к блоку «Защита»: тут на карте отмечены стейкхолдеры, для которых требуются специальные инициативы по защите их интересов, потому что для них проект достаточно важен, а их влияние на его реализацию незначительно (либо они вообще не имеют возможности повлиять).'
            return result


# making objects for analys
analys_param_1 = impact_level()
analys_param_2 = engagement()
analys_param_3 = needs_and_requirements()
analys_param_4 = expectations()
analys_param_5 = level_of_interests()
analys_param_6 = effect()
analys_param_7 = problematical_character()
analys_param_8 = communicativity()

# making object to become a result
final_result = result()


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


Logo = resource_path("Logo.png")


class Stakeholder(QtWidgets.QMainWindow, Ui_Dialog):
    '''
    This class will have all fields and methods for stakeholder analys
    '''

    __result = []

    def __init__(self):
        '''
        Konstruktor of our class
        '''
        super().__init__()

        # Создание формы и Ui (наш дизайн)
        self.setupUi(self)

        # Показать наше окно
        self.show()

        self.pushButton.clicked.connect(self.making_analys)

    def get_data(self):
        '''
        Method, thar return some data of person
        '''

        return self.__result

    def making_analys(self):
        '''
        This function will make analys and get the result
        '''
        # score result of parametrs
        name_line = self.lineEdit.text()
        family_name_line = self.lineEdit_2.text()
        organisation_name_line = self.lineEdit_3.text()
        position_line = self.lineEdit_4.text()
        contacts_line = self.lineEdit_5.text()

        # data that we become from comboboxes(our main-parametrs for analys)
        param_1 = self.comboBox.currentText()
        param_2 = self.comboBox_2.currentText()
        param_3 = self.comboBox_3.currentText()
        param_4 = self.comboBox_4.currentText()
        param_5 = self.comboBox_5.currentText()
        param_6 = self.comboBox_6.currentText()
        param_7 = self.comboBox_7.currentText()
        param_8 = self.comboBox_8.currentText()

        self.__result.append(name_line)
        self.__result.append(family_name_line)
        self.__result.append(organisation_name_line)
        self.__result.append(position_line)
        self.__result.append(contacts_line)
        self.__result.append(param_1)
        self.__result.append(param_2)
        self.__result.append(param_3)
        self.__result.append(param_4)
        self.__result.append(param_5)
        self.__result.append(param_6)
        self.__result.append(param_7)
        self.__result.append(param_8)
        g_1 = analys_param_1.make_analys(self.__result[5])
        g_2 = analys_param_2.make_analys(self.__result[6])
        g_3 = analys_param_3.make_analys(self.__result[7])
        g_4 = analys_param_4.make_analys(self.__result[8])
        g_5 = analys_param_5.make_analys(self.__result[9])
        g_6 = analys_param_6.make_analys(self.__result[10])
        g_7 = analys_param_7.make_analys(self.__result[11])
        g_8 = analys_param_8.make_analys(self.__result[12])

        # final result for matrix of stakeholder
        result_x = final_result.scoring_1(g_1=g_1, g_2=g_2, g_7=g_7, g_8=g_8)
        result_y = final_result.scoring_2(g_3=g_3, g_4=g_4, g_5=g_5, g_6=g_6)

        # what to do with this person
        final = final_result.make_simple_analys(x=result_x, y=result_y)
        self.__result.append(final)

        if final == 'Данная личность не является стейкхолдером.':
            final_data = os.startfile(r'.\not.png')
        elif final == 'Данный стейкхолдер относится к блоку «Хорошие отношения»:\nC этими стейкхолдерами необходимо установить тесные рабочие отношения, потому что для них проект важен, они вовлечены в реализацию и активно влияют на процесс и результат.':
            final_data = os.startfile(r'.\good_relation.png')
        elif final == 'Данный стейкхолдер относится к блоку «Мониторинг»:\nОни имеют власть над реализацией проекта, но не слишком заинтересованы в нем.\nПри таком сочетании факторов они могут стать источниками рисков, поэтому необходимы тщательный мониторинг и внимательный менеджмент.':
            final_data = os.startfile(r'.\monitor.png')
        elif final == 'Данный стейкхолдер относится к блоку «Низкий приоритет»: этот стейкхолдер вовлечен и относительно заинтересован, но от него зависит не так много, поэтому с точки зрения распределения менеджерского внимания, у него низкий приоритет.':
            final_data = os.startfile(r'.\low_priority.png')
        else:
            final_data = os.startfile(r'.\protect.png')

        if os.path.exists(r'.\database.xlsx') == True:
            # list of strings

            # Calling DataFrame constructor on list
            # df = pd.DataFrame(arr)

            wb = load_workbook('database.xlsx')
            ws = wb.active

            # for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(self.__result)

            wb.save('database.xlsx')
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(['Имя', 'Фамилия', 'Организация', 'Должность', 'Контактные данные', 'Уровень влияния',
                       'Заинтересованность', 'Нужды и потребности', 'Ожидания', 'Уровень заинтересованности',
                       'Эффект на ПМ', 'Проблематичность', 'Коммуникативность', 'Советы'])
            ws.append(self.__result)
            wb.save("database.xlsx")

        self.__result = []

        return final_data


if __name__ == "__main__":
    # Новый экземпляр QApplication
    app = QtWidgets.QApplication(sys.argv)

    # Сoздание инстанса класса, который мы создадим далее
    stakeholder = Stakeholder()
    # stakeholder.get_data()
    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog)

    sys.exit(app.exec_())
